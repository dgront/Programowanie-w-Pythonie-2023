from openpyxl import Workbook
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill, Font
from colormap import rgb2hex

N = 20

w = Workbook()
arkusz = w.active
for row in range(N):
    red = int(255 / N * row)
    for col in range(N):
        blue = int(255 / N * col)
        arkusz.cell(row+1, col+1, (row+1) * (col+1))
        color_tla_as_str = rgb2hex(red, 150, blue)
        color_txt_as_str = rgb2hex(blue, 10, red)
        komorka = arkusz.cell(row+1, col+1)
        color_tla = Color(color_tla_as_str[1:], type='rgb')
        color_txt = Color(color_txt_as_str[1:], type='rgb')
        komorka.fill = PatternFill(patternType='solid', fill_type='solid', fgColor=color_tla)
        komorka.font = Font(color=color_txt_as_str[1:])
w.save("plik2.xlsx")
